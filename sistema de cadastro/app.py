from typing import Tuple
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
import os

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()  
        self.todo_sistema()  
            
    def layout_config(self):
        self.title("Gestão de Cadastro de Clientes")
        self.geometry("700x500")
        
    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", text_color=['#000', "#fff"])
        self.lb_apm.place(x=50, y=430) 
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm)
        self.opt_apm.place(x=50, y=460) 

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=900, height=50, corner_radius=0, fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Gestão de Cadastro de Clientes", font=("Century Gothic bold", 24), text_color="#fff")
        title.place(x=200, y=10)  
        
        span = ctk.CTkLabel(self, text="Por favor, preencha todos os campos do formulário!", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        span.place(x=50, y=70)
        
        def submit():
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            address = address_value.get()
            obs = obs_entry.get(1.0, END)
            
            if (name == "" or contact == "" or age == "" or address == ""):
                messagebox.showerror("Sistema", "ERRO!\nPor favor preencha todos os dados!")
            else:
                try:
                    # Verificar se o arquivo existe
                    if not os.path.exists('Clientes.xlsx'):
                        # Se não existir, criar um novo arquivo
                        ficheiro = openpyxl.Workbook()
                        folha = ficheiro.active
                        # Criar cabeçalhos, se necessário
                        folha.append(["Nome", "Contato", "Idade", "Gênero", "Endereço", "Observação"])
                    else:
                        # Carregar o arquivo Excel existente
                        ficheiro = openpyxl.load_workbook('Clientes.xlsx')
                        folha = ficheiro.active

                    # Determinar a próxima linha disponível
                    proxima_linha = folha.max_row + 1

                    # Adicionar os valores nas células correspondentes na próxima linha
                    folha.cell(column=1, row=proxima_linha, value=name)
                    folha.cell(column=2, row=proxima_linha, value=contact)
                    folha.cell(column=3, row=proxima_linha, value=age)
                    folha.cell(column=4, row=proxima_linha, value=gender)
                    folha.cell(column=5, row=proxima_linha, value=address)
                    folha.cell(column=6, row=proxima_linha, value=obs)

                    # Salvar as alterações no arquivo Excel
                    ficheiro.save('Clientes.xlsx')
                    messagebox.showinfo("Sistema", "Dados salvos com sucesso!")

                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao salvar os dados: {e}")
        
        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            gender_combobox.set("Masculino")
            obs_entry.delete(1.0, END)
    
        name_value = StringVar() 
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()
        
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=("Century Gothic bold", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century Gothic bold", 16), fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200, textvariable=address_value, font=("Century Gothic bold", 16), fg_color="transparent")
        
        gender_combobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino"], font=("Century Gothic bold", 14))
        gender_combobox.set("Masculino")  
        
        obs_entry = ctk.CTkTextbox(self, width=500, height=150, font=("Arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")
        
        # Labels 
        lb_name = ctk.CTkLabel(self, text="Nome completo:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(self, text="Contato:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_age = ctk.CTkLabel(self, text="Idade:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_gender = ctk.CTkLabel(self, text="Gênero:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_address = ctk.CTkLabel(self, text="Endereço:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observação:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        btn_submit = ctk.CTkButton(self, text="Adicionar dados".upper(), command=submit, fg_color="#151", hover_color="#131")
        btn_submit.place(x=300, y=420)
        
        btn_clear = ctk.CTkButton(self, text="Limpar dados".upper(), command=clear, fg_color="#555", hover_color="#131")
        btn_clear.place(x=500, y=420)
        
        # Posicionamento das Labels e Entradas
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)
        
        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)
        
        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)
        
        lb_gender.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)
        
        lb_address.place(x=50, y=190)
        address_entry.place(x=50, y=220)
        
        lb_obs.place(x=50, y=260)
        obs_entry.place(x=150, y=260)
        
    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

if __name__ == "__main__":
    app = App()
    app.mainloop()
